# 영문 xlsx 파일 누락 수정 코드

import xlrd
from openpyxl import load_workbook

rb_ko = xlrd.open_workbook('../xlsx/개설교과목정보(한글).xlsx')
sh_ko = rb_ko.sheet_by_index(0)

rb_en = xlrd.open_workbook('../xlsx/개설교과목정보(영문).xlsx')
sh_en = rb_en.sheet_by_index(0)

empty_department_index = []

interdisciplinary_programs_ko = ["지식융합미디어학부", "교육문화연계전공", "공공인재 연계전공", "여성학 연계전공", "정치학/경제학/철학 연계전공", "스포츠미디어 연계전공", "바이오융합기술 연계전공", "스타트업연계전공", "융합소프트웨어연계전공", "한국발전과국제개발협력연계전공", "빅데이터사이언스연계전공", "인공지능연계전공", "한국사회문화연계전공"]
interdisciplinary_programs_en = ["School of Media, Arts and Science", "Educational Culture", "Public Leadership", "Gender Studies", "PEP(Political Science, Economics and Philosophy", "Sports Media", "Department of Biomedical Engineering", "Sogang Startup Academy", "Samsung Convergence Software Course", "Development of Korea and International Cooperation Development", "Big Data Science", "Artificial Intelligence", "Korean Society and Culture"]

for row_num in range(1, sh_en.nrows):
    row_values_ko = sh_ko.row_values(row_num)
    row_values_en = sh_en.row_values(row_num)

    if row_values_en[3] == '':
        empty_department_index.append(row_num)

wb = load_workbook('../xlsx/개설교과목정보(영문).xlsx')
ws = wb['개설교과목정보(영문)']

for i in range(len(empty_department_index)):
    index = 'D'+str(empty_department_index[i]+1)
    text = interdisciplinary_programs_en[interdisciplinary_programs_ko.index(sh_ko.row_values(empty_department_index[i])[3])]

    ws[index] = text

wb.save('../xlsx/개설교과목정보(영문)_누락수정.xlsx')